# ============================================================
# excel_service.py
# Servicio para generar el archivo Excel de la proforma
# ============================================================

import os       # Para manejar rutas y carpetas
import copy     # Para copiar el formato de las celdas
from openpyxl import load_workbook  # Para leer y modificar Excel

# Ruta de la plantilla base (debe estar en la carpeta AppProforma)
TEMPLATE_PATH = "plantilla.xlsx"

# Carpeta donde se guardarán las proformas generadas
OUTPUT_DIR = "proformas"


def crear_carpeta_proformas():
    """Crea la carpeta proformas si no existe"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


def copiar_formato_fila(ws, fila_origen: int, fila_destino: int):
    """Copia el formato completo de una fila a otra para mantener bordes y estilos"""
    for col in range(1, 9):  # Columnas A hasta H
        src = ws.cell(row=fila_origen, column=col)   # Celda origen
        dst = ws.cell(row=fila_destino, column=col)  # Celda destino

        if src.has_style:  # Solo copiar si la celda tiene algún estilo
            dst.font = copy.copy(src.font)            # Copiar fuente
            dst.fill = copy.copy(src.fill)            # Copiar relleno/color
            dst.border = copy.copy(src.border)        # Copiar bordes
            dst.alignment = copy.copy(src.alignment)  # Copiar alineación
            dst.number_format = src.number_format     # Copiar formato numérico


def generar_excel(proforma: dict, nombre_archivo: str,
                  tipo_descuento: str = "ninguno",
                  porcentaje_descuento: float = 0) -> str:
    """
    Genera el Excel de la proforma basado en la plantilla.

    Args:
        proforma: Diccionario completo con datos de la proforma
        nombre_archivo: Nombre sin extensión del archivo a generar
        tipo_descuento: 'ninguno', 'individual' o 'global'
        valor_descuento_global: Valor en $ del descuento global

    Returns:
        Ruta completa del archivo Excel generado
    """
    crear_carpeta_proformas()         # Asegurar que existe la carpeta
    wb = load_workbook(TEMPLATE_PATH) # Cargar la plantilla Excel
    ws = wb.active                    # Activar la hoja de trabajo

    # =========================================================
    # SECCIÓN 1: DATOS DEL CLIENTE
    # =========================================================
    ws['B10'] = proforma['cliente']['nombre']  # Nombre del cliente
    ws['F10'] = proforma['fecha']              # Fecha de la proforma
    ws['B11'] = proforma['cliente']['lugar']   # Dirección o lugar
    ws['F11'] = '-'                            # MODELO (no utilizado)
    ws['B12'] = proforma['cliente']['obra']    # Nombre de la obra

    # =========================================================
    # SECCIÓN 2: TABLA DE PRODUCTOS
    # =========================================================
    productos = proforma['productos']  # Lista de productos
    N = len(productos)                 # Cantidad de productos
    PRIMERA_FILA = 17                  # Primera fila de productos en la plantilla

    # La plantilla tiene 2 filas de productos (17 y 18)
    # Para cada producto extra (3, 4, 5...) insertamos una fila
    for i in range(max(0, N - 2)):
        fila_insertar = PRIMERA_FILA + 2 + i   # Posición de inserción
        ws.insert_rows(fila_insertar)            # Insertar fila vacía
        # Copiar formato de la fila anterior para mantener bordes
        copiar_formato_fila(ws, fila_insertar - 1, fila_insertar)

    # Llenar los datos de cada producto
    for i, prod in enumerate(productos):
        fila = PRIMERA_FILA + i  # Fila actual del producto

        ws.cell(row=fila, column=1).value = prod['numero']      # A: Nº consecutivo
        ws.cell(row=fila, column=2).value = prod['producto']    # B: Nombre
        ws.cell(row=fila, column=3).value = prod['descripcion'] # C: Descripción
        # D: Descripción adicional (si no hay, coloca "-")
        ws.cell(row=fila, column=4).value = prod.get('descripcion_adicional', '-')
        ws.cell(row=fila, column=5).value = prod['cantidad']          # E: Cantidad
        ws.cell(row=fila, column=6).value = prod['precio_unitario']   # F: Precio

        # G: Descuento individual en valor absoluto ($)
        if tipo_descuento == "individual":
            ws.cell(row=fila, column=7).value = prod.get('descuento', 0)
        else:
            ws.cell(row=fila, column=7).value = 0

        # H: Total = (Cantidad × Precio) - Descuento individual
        ws.cell(row=fila, column=8).value = f'=((E{fila}*F{fila})-G{fila})'

    # =========================================================
    # CALCULAR FILAS DINÁMICAS
    # Cada producto extra desplaza todas las filas siguientes
    # =========================================================
    desplazamiento = max(0, N - 2)  # Filas insertadas extras

    fila_placeholder = 19 + desplazamiento  # Fila placeholder (se limpiará)
    fila_subtotal    = 21 + desplazamiento  # Fila SUBTOTAL
    fila_desc_global = 22 + desplazamiento  # Fila DESCUENTO GLOBAL
    fila_iva         = 23 + desplazamiento  # Fila IVA
    fila_total       = 24 + desplazamiento  # Fila TOTAL FINAL
    fila_nota        = 25 + desplazamiento  # Fila para NOTA (actualmente vacía)
    fila_a1          = 27 + desplazamiento  # Fila Anticipo 1
    fila_a2          = 28 + desplazamiento  # Fila Anticipo 2
    fila_a3          = 29 + desplazamiento  # Fila Anticipo 3
    fila_total_proy  = 30 + desplazamiento  # Fila TOTAL PROYECTO / SALDO

    # Limpiar filas de productos no utilizados (placeholder y sobrantes)
    for fila_limpiar in range(PRIMERA_FILA + N, fila_placeholder + 1):
        for col in range(1, 9):
            ws.cell(row=fila_limpiar, column=col).value = None

    # =========================================================
    # SECCIÓN 3: SUBTOTAL
    # =========================================================
    ultima_fila_prod = PRIMERA_FILA + N - 1  # Última fila con producto real
    # Sumar solo las filas reales de productos
    ws.cell(row=fila_subtotal, column=8).value = \
        f'=SUM(H{PRIMERA_FILA}:H{ultima_fila_prod})'

    # =========================================================
    # SECCIÓN 4: DESCUENTOS
    # El porcentaje ingresado por el bot se convierte en
    # una fórmula que calcula el valor en $ directamente en Excel
    # =========================================================
    pct = porcentaje_descuento / 100  # Convertir % a decimal (ej: 5 → 0.05)

    if tipo_descuento == "ninguno":
        # Sin descuento: ocultar columna G y poner 0 en descuento global
        ws.column_dimensions['G'].hidden = True
        ws.cell(row=fila_desc_global, column=8).value = 0

    elif tipo_descuento == "individual":
        # Descuento individual: mostrar columna G
        # Cada celda G calcula: (Cantidad × Precio) × porcentaje
        ws.column_dimensions['G'].hidden = False
        for i in range(N):
            fila = PRIMERA_FILA + i
            # Fórmula: G = E×F×porcentaje → da el valor en $ del descuento
            ws.cell(row=fila, column=7).value = f'=E{fila}*F{fila}*{pct}'
        # Descuento global en 0 (no aplica)
        ws.cell(row=fila_desc_global, column=8).value = 0

    elif tipo_descuento == "global":
        # Descuento global: ocultar columna G
        # H22 calcula: Subtotal × porcentaje → da el valor en $ del descuento
        ws.column_dimensions['G'].hidden = True
        ws.cell(row=fila_desc_global, column=8).value = \
            f'=H{fila_subtotal}*{pct}'
        
    # =========================================================
    # SECCIÓN 5: IVA
    # =========================================================
    if proforma['incluye_iva']:
        # Con IVA: calcular 15% sobre (Subtotal - Descuento Global)
        ws.cell(row=fila_iva, column=8).value = \
            f'=(H{fila_subtotal}-H{fila_desc_global})*0.15'
        ws.cell(row=fila_iva, column=6).value = 'I.V.A. 15%:'
    else:
        # Sin IVA: colocar 0 y actualizar etiqueta
        ws.cell(row=fila_iva, column=8).value = 0
        ws.cell(row=fila_iva, column=6).value = 'I.V.A. 0%:'

    # Total final: (Subtotal - Descuento Global) + IVA
    ws.cell(row=fila_total, column=8).value = \
        f'=(H{fila_subtotal}-H{fila_desc_global})+H{fila_iva}'

    # =========================================================
    # SECCIÓN 6: NOTA O CONDICIÓN ESPECIAL
    # Va entre el total y la tabla de anticipos
    # =========================================================
    if proforma.get('nota') and proforma['nota'].strip():
        ws.insert_rows(fila_nota)  # Insertar una fila para la nota
        ws.cell(row=fila_nota, column=2).value = f"NOTA: {proforma['nota']}"
        # Los anticipos se desplazan una fila más
        fila_a1 += 1
        fila_a2 += 1
        fila_a3 += 1
        fila_total_proy += 1

    # =========================================================
    # SECCIÓN 7: TABLA DE ANTICIPOS
    # =========================================================
    anticipos = proforma['anticipos']
    pct1 = anticipos['anticipo_1']['porcentaje'] / 100  # Ej: 0.60 para 60%
    pct2 = anticipos['anticipo_2']['porcentaje'] / 100  # Ej: 0.20 para 20%

    # Anticipo 1: porcentaje del total
    ws.cell(row=fila_a1, column=4).value = f'=H{fila_total}*{pct1}'

    # Anticipo 2: porcentaje del total
    ws.cell(row=fila_a2, column=4).value = f'=H{fila_total}*{pct2}'

    # Anticipo 3: lo que queda del total menos los 2 primeros anticipos
    ws.cell(row=fila_a3, column=4).value = \
        f'=H{fila_total}-(D{fila_a1}+D{fila_a2})'

    # Total del proyecto: suma de los 3 anticipos
    ws.cell(row=fila_total_proy, column=4).value = \
        f'=SUM(D{fila_a1}:D{fila_a3})'

    # Saldo a pagar: total menos pagos ya realizados (F27-F29)
    ws.cell(row=fila_total_proy, column=6).value = \
        f'=D{fila_total_proy}-SUM(F{fila_a1}:F{fila_a3})'

    # =========================================================
    # GUARDAR EL ARCHIVO EXCEL
    # =========================================================
    ruta_archivo = f"{OUTPUT_DIR}/{nombre_archivo}.xlsx"
    wb.save(ruta_archivo)  # Guardar en la carpeta proformas

    return ruta_archivo  # Retornar ruta para enviar por WhatsApp